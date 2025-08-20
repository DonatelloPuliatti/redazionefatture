from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime
import numpy as np

percorso_chromedriver = "C:/webdriver/chromedriver.exe"  # Adatta se serve
service = Service(percorso_chromedriver)
driver = webdriver.Chrome(service=service)
wait = WebDriverWait(driver, 15)

tutti_dati = []  # qui mettiamo tutte le righe

for pagina in range(1, 9):
    url = f"https://www.borsaitaliana.it/borsa/obbligazioni/mot/btp/lista.html?&page={pagina}"
    driver.get(url)



    righe = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

    for riga in righe:
        celle = riga.find_elements(By.TAG_NAME, "td")
        dati = [c.text for c in celle]
        if dati:
            print(dati)
            tutti_dati.append(dati)  # aggiungi i dati

driver.quit()

colonne = ['ISIN', 'Nome', 'Prezzo', 'Cedola %', 'Scadenza', 'Altro']

df = pd.DataFrame(tutti_dati, columns=colonne)

df['ISIN'] = df['ISIN'].astype(str)
df['Nome'] = df['Nome'].astype(str)

df['Prezzo'] = df['Prezzo'].str.replace(',', '.')
df['Prezzo'] = pd.to_numeric(df['Prezzo'], errors='coerce')

df['Cedola %'] = df['Cedola %'].str.replace(',', '.')
df['Cedola %'] = pd.to_numeric(df['Cedola %'], errors='coerce')

df['Scadenza'] = pd.to_datetime(df['Scadenza'], dayfirst=True, errors='coerce')

# Se vuoi togliere l'ora (e mantenere datetime)
df['Scadenza'] = df['Scadenza'].dt.normalize()

# Se invece vuoi la stringa formattata "gg-mm-aa"


df['Cedola annuale %'] = df['Cedola %'] * 2
df['Prezzo valore nominale € 10 K'] = df['Prezzo'] * 100
df['Prezzo valore nominale € 10 K'] = df['Prezzo'] * 100
df['Cedola semestrale su € 10 K'] = df['Cedola %'] * 100
df['Cedola semestrale su € 10 K con ritenuta'] = df['Cedola semestrale su € 10 K'] * 0.875
df['Cedola annuale su € 10 K'] = df['Cedola annuale %'] * 100
df['Cedola annuale su € 10 K con ritenuta'] = df['Cedola annuale su € 10 K'] * 0.875
df['Rendimento attuale %'] = df['Cedola annuale %'] / df['Prezzo'] * 100
df['Rendimento attuale %'] = np.trunc(df['Rendimento attuale %'] * 1000) / 1000
oggi = pd.to_datetime(datetime.today()).normalize()
df['Scadenza numerica'] = (df['Scadenza'] - oggi).dt.days
df['Scadenza'] = df['Scadenza'].dt.strftime('%d-%m-%y')
df['Numero cedole'] = df['Scadenza numerica']//182.5 + 1
df['Ricavo totale'] = df['Numero cedole'] * df['Cedola %'] + 100
df['Guadagno totale lordo'] = df['Ricavo totale'] - df['Prezzo']
df['Guadagno totale netto'] = df['Guadagno totale lordo'] * 0.875
df['Guadagno totale netto'] = np.trunc(df['Guadagno totale netto'] * 1000) / 1000
df['Guadagno medio lordo'] = df['Guadagno totale lordo'] / df['Scadenza numerica'] * 365
df['Guadagno medio lordo'] = np.trunc(df['Guadagno medio lordo'] * 1000) / 1000
df['Guadagno medio netto'] = df['Guadagno medio lordo'] * 0.875
df['Guadagno medio netto'] = np.trunc(df['Guadagno medio netto'] * 1000) / 1000
colonne_da_esportare = ['ISIN', 'Nome', 'Prezzo', 'Cedola %', 'Scadenza', 'Cedola annuale %', 'Rendimento attuale %', 'Numero cedole', 'Ricavo totale', 'Guadagno totale lordo', 'Guadagno totale netto', 'Guadagno medio lordo', 'Guadagno medio netto']



# Eliminiamo la colonna Altro
df = df.drop(columns=['Altro'])

print(df.dtypes)

df[colonne_da_esportare].to_excel("btp_dati.xlsx", index=False)

wb = load_workbook("btp_dati.xlsx")
ws = wb.active

# Raddoppia la larghezza delle colonne in base al contenuto attuale
for col_idx, col_cells in enumerate(ws.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)
    for cell in col_cells:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    # Imposta la larghezza raddoppiata (moltiplica per 2)
    ws.column_dimensions[col_letter].width = max_length * 1.2

# Salva il file modificato
wb.save("btp_dati.xlsx")