# borsa.py
import os
import io
import sys
import json
import zipfile
import logging
from datetime import datetime
from bs4 import BeautifulSoup
import numpy as np
import pandas as pd
from flask import request, render_template
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import SessionNotCreatedException, WebDriverException

# webdriver_manager resta come fallback finale per il driver
from webdriver_manager.chrome import ChromeDriverManager

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("borsa")


# ----------------------- UTIL -----------------------
def _first_existing(paths):
    for p in paths:
        if p and os.path.exists(p):
            return p
    return None

def _find_file_recursive(root_dir, filename):
    """Cerca ricorsivamente 'filename' dentro root_dir. Ritorna il primo path trovato o None."""
    for base, _, files in os.walk(root_dir):
        if filename in files:
            return os.path.join(base, filename)
    return None

def _ensure_executable(path):
    try:
        mode = os.stat(path).st_mode
        # aggiunge permesso esecuzione per utente se mancante
        os.chmod(path, mode | 0o100)
    except Exception:
        pass

def _log_tree(root_dir, depth=2):
    """Stampa (nei log) una piccola vista ad albero utile al debug."""
    try:
        for base, dirs, files in os.walk(root_dir):
            level = base.replace(root_dir, "").count(os.sep)
            if level > depth:
                continue
            indent = "  " * level
            logger.info("%s%s/", indent, os.path.basename(base) or root_dir)
            for f in files[:20]:
                logger.info("%s  %s", indent, f)
    except Exception:
        pass


def _ensure_portable_chrome(base_dir="/tmp/cft"):
    """
    Scarica Chrome for Testing (browser + chromedriver) portabili (senza root) se non già presenti.
    Restituisce (chrome_binary_path, chromedriver_path).
    Le zip CfT si estraggono in cartelle tipo:
      - chrome-linux64/chrome
      - chromedriver-linux64/chromedriver
    """
    import requests  # assicurati: presente in requirements.txt

    os.makedirs(base_dir, exist_ok=True)
    chrome_bin = _find_file_recursive(base_dir, "chrome")
    driver_bin = _find_file_recursive(base_dir, "chromedriver")

    if chrome_bin and driver_bin:
        logger.info(f"Chrome portabile già presente: {chrome_bin}")
        logger.info(f"Chromedriver portabile già presente: {driver_bin}")
        _ensure_executable(chrome_bin)
        _ensure_executable(driver_bin)
        return chrome_bin, driver_bin

    index_url = "https://googlechromelabs.github.io/chrome-for-testing/known-good-versions-with-downloads.json"
    logger.info("Scarico indice Chrome for Testing...")
    r = requests.get(index_url, timeout=45)
    r.raise_for_status()
    data = r.json()

    entries = data.get("versions", [])
    if not entries:
        raise RuntimeError("Indice Chrome for Testing vuoto.")

    latest = entries[-1]
    version = latest["version"]

    def _find_url(pkg_key):
        for item in latest["downloads"].get(pkg_key, []):
            if item.get("platform") == "linux64":
                return item["url"]
        return None

    chrome_url = _find_url("chrome")
    driver_url = _find_url("chromedriver")
    if not chrome_url or not driver_url:
        raise RuntimeError("URL Chrome/Chromedriver linux64 non trovati nell'indice CfT.")

    def _download_and_extract(url, target_dir):
        os.makedirs(target_dir, exist_ok=True)
        logger.info(f"Scarico: {url}")
        resp = requests.get(url, timeout=180)
        resp.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
            zf.extractall(target_dir)

    _download_and_extract(chrome_url, base_dir)
    _download_and_extract(driver_url, base_dir)

    # dopo l’estrazione, cerchiamo i binari col nome corretto
    _log_tree(base_dir, depth=2)
    chrome_bin = _find_file_recursive(base_dir, "chrome")
    driver_bin = _find_file_recursive(base_dir, "chromedriver")

    if not chrome_bin:
        raise RuntimeError("Chrome portabile non trovato dopo l'estrazione.")
    if not driver_bin:
        raise RuntimeError("Chromedriver portabile non trovato dopo l'estrazione.")

    _ensure_executable(chrome_bin)
    _ensure_executable(driver_bin)

    logger.info(f"Chrome for Testing pronto (v{version}) in {chrome_bin}")
    logger.info(f"Chromedriver abbinato in {driver_bin}")
    return chrome_bin, driver_bin


# ----------------------- SELENIUM -----------------------
def make_driver():
    """
    Avvia un Chrome headless in modo robusto:
      A) Usa browser di sistema se presente (CHROME_BIN/GOOGLE_CHROME_BIN o percorsi noti).
      B) Altrimenti scarica Chrome for Testing portabile e usa il chromedriver corrispondente.
      C) Ultimo fallback: webdriver_manager solo per il driver (se esiste un browser di sistema).
    """
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=it-IT")
    options.add_argument(
        "--user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )

    chrome_candidates = [
        os.getenv("CHROME_BIN"),
        os.getenv("GOOGLE_CHROME_BIN"),
        "/usr/bin/chromium-browser",
        "/usr/bin/chromium",
        "/usr/bin/google-chrome",
        "/snap/bin/chromium",
    ]
    driver_candidates = [
        os.getenv("CHROMEDRIVER_PATH"),
        "/usr/bin/chromedriver",
        "/usr/local/bin/chromedriver",
    ]

    chrome_path = _first_existing(chrome_candidates)
    sys_driver = _first_existing(driver_candidates)

    # Caso A: browser di sistema presente
    if chrome_path:
        logger.info(f"Trovato browser di sistema: {chrome_path}")
        _ensure_executable(chrome_path)
        options.binary_location = chrome_path
        try:
            if sys_driver:
                logger.info(f"Uso chromedriver di sistema: {sys_driver}")
                _ensure_executable(sys_driver)
                service = Service(executable_path=sys_driver)
                return webdriver.Chrome(service=service, options=options)
            else:
                logger.info("Chromedriver di sistema non trovato: uso webdriver_manager per il driver...")
                service = Service(ChromeDriverManager().install())
                return webdriver.Chrome(service=service, options=options)
        except (SessionNotCreatedException, WebDriverException) as e:
            logger.warning(f"Avvio con browser di sistema fallito: {e}. Provo Chrome portabile...")

    # Caso B: Chrome for Testing portabile
    try:
        chrome_portable, driver_portable = _ensure_portable_chrome("/tmp/cft")
        options.binary_location = chrome_portable
        service = Service(executable_path=driver_portable)
        logger.info("Avvio Chrome for Testing portabile...")
        return webdriver.Chrome(service=service, options=options)
    except Exception as e:
        logger.exception(f"Fallback CfT portabile fallito: {e}")

    # Caso C: fallback estremo (solo se esiste un browser nel PATH esterno)
    try:
        logger.info("Ultimo tentativo: webdriver_manager con binary di default...")
        service = Service(ChromeDriverManager().install())
        return webdriver.Chrome(service=service, options=options)
    except Exception as e:
        msg = f"Impossibile avviare Chrome (tutti i tentativi falliti): {e}"
        logger.exception(msg)
        raise RuntimeError(msg)


# ----------------------- ROUTE -----------------------
def borsa():
    if request.method == "GET":
        return render_template("borsa.html")

    tipologia = request.form.get("tipologia")
    if tipologia != "BTP":
        return render_template(
            "borsa.html",
            risultato="Seleziona BTP per eseguire l'elaborazione.",
            tabella=None
        )

    driver = None
    try:
        driver = make_driver()

        tutti_dati = []
        # Pagine BTP su Borsa Italiana (1..8)
        for pagina in range(1, 9):
            url = f"https://www.borsaitaliana.it/borsa/obbligazioni/mot/btp/lista.html?&page={pagina}"
            logger.info(f"Scarico: {url}")
            driver.get(url)

            html = driver.page_source
            soup = BeautifulSoup(html, "lxml")
            for tr in soup.select("table tbody tr"):
                celle = [td.get_text(strip=True) for td in tr.find_all("td")]
                if celle:
                    tutti_dati.append(celle)

        # Colonne attese dalla tabella corrente
        colonne = ["ISIN", "Nome", "Prezzo", "Cedola %", "Scadenza", "Altro"]
        df = pd.DataFrame(tutti_dati, columns=colonne)

        # Tipi
        df["ISIN"] = df["ISIN"].astype(str)
        df["Nome"] = df["Nome"].astype(str)

        # Prezzo e Cedola
        df["Prezzo"] = (
            df["Prezzo"]
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df["Prezzo"] = pd.to_numeric(df["Prezzo"], errors="coerce")

        df["Cedola %"] = (
            df["Cedola %"]
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df["Cedola %"] = pd.to_numeric(df["Cedola %"], errors="coerce")

        # Scadenza
        df["Scadenza"] = pd.to_datetime(df["Scadenza"], dayfirst=True, errors="coerce").dt.normalize()

        # Calcoli
        df["Cedola annuale %"] = df["Cedola %"] * 2
        df["Prezzo valore nominale € 10 K"] = df["Prezzo"] * 100
        df["Cedola semestrale su € 10 K"] = df["Cedola %"] * 100
        df["Cedola semestrale su € 10 K con ritenuta"] = df["Cedola semestrale su € 10 K"] * 0.875
        df["Cedola annuale su € 10 K"] = df["Cedola annuale %"] * 100
        df["Cedola annuale su € 10 K con ritenuta"] = df["Cedola annuale su € 10 K"] * 0.875
        df["Rendimento attuale %"] = df["Cedola annuale %"] / df["Prezzo"] * 100
        df["Rendimento attuale %"] = np.trunc(df["Rendimento attuale %"] * 1000) / 1000

        oggi = pd.to_datetime(datetime.today()).normalize()
        df["Scadenza numerica"] = (df["Scadenza"] - oggi).dt.days
        df["Scadenza"] = df["Scadenza"].dt.strftime("%d-%m-%y")

        df["Numero cedole"] = df["Scadenza numerica"] // 182.5 + 1
        df["Ricavo totale"] = df["Numero cedole"] * df["Cedola %"] + 100
        df["Guadagno totale lordo"] = df["Ricavo totale"] - df["Prezzo"]
        df["Guadagno totale netto"] = df["Guadagno totale lordo"] * 0.875
        df["Guadagno totale netto"] = np.trunc(df["Guadagno totale netto"] * 1000) / 1000
        df["Guadagno medio lordo"] = df["Guadagno totale lordo"] / df["Scadenza numerica"] * 365
        df["Guadagno medio lordo"] = np.trunc(df["Guadagno medio lordo"] * 1000) / 1000
        df["Guadagno medio netto"] = df["Guadagno medio lordo"] * 0.875
        df["Guadagno medio netto"] = np.trunc(df["Guadagno medio netto"] * 1000) / 1000

        # Elimina "Altro" se presente
        if "Altro" in df.columns:
            df = df.drop(columns=["Altro"])

        colonne_da_esportare = [
            "ISIN",
            "Nome",
            "Prezzo",
            "Cedola %",
            "Scadenza",
            "Cedola annuale %",
            "Rendimento attuale %",
            "Numero cedole",
            "Ricavo totale",
            "Guadagno totale lordo",
            "Guadagno totale netto",
            "Guadagno medio lordo",
            "Guadagno medio netto",
        ]

        # Salvataggio XLSX
        output_path = "/tmp/btp_dati.xlsx" if os.access("/tmp", os.W_OK) else "btp_dati.xlsx"
        df[colonne_da_esportare].to_excel(output_path, index=False)

        # Auto-larghezza colonne
        wb = load_workbook(output_path)
        ws = wb.active
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(12, int(max_length * 1.2))
        wb.save(output_path)

        # Tabella HTML
        tabella_html = df.to_html(
            classes="tabella-risultati display nowrap", table_id="btpTable", index=False
        )

        return render_template(
            "borsa.html",
            risultato=f"Elaborazione {tipologia} completata. File salvato come {os.path.basename(output_path)}",
            tabella=tabella_html,
        )

    except Exception as e:
        logger.exception("Errore durante scraping/elaborazione BTP")
        return render_template(
            "borsa.html",
            risultato=f"Si è verificato un errore: {type(e).__name__} - {e}",
            tabella=None
        ), 500

    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass

