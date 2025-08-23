# borsa.py
import os
import logging
from datetime import datetime
from bs4 import BeautifulSoup
import numpy as np
import pandas as pd
from flask import request, render_template
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import SessionNotCreatedException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager

logging.basicConfig(level=logging.INFO)


# ---------- Selenium helper: compatibile Windows locale + Render (Linux) ----------
def _find_first_existing(paths):
    """Ritorna il primo path esistente nella lista, altrimenti None."""
    for p in paths:
        if p and os.path.exists(p):
            return p
    return None



def make_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=it-IT")
    options.add_argument(
        "--user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )

    # 1) prova con i binari di sistema (Render + Aptfile)
    chrome_candidates = [
        os.getenv("CHROME_BIN"),
        os.getenv("GOOGLE_CHROME_BIN"),
        "/usr/bin/chromium-browser",
        "/usr/bin/chromium",
        "/usr/bin/google-chrome",
        "/snap/bin/chromium",
    ]
    driver_candidates = [
        os.getenv("CHROMEDRIVER_PATH"),
        "/usr/bin/chromedriver",
        "/usr/local/bin/chromedriver",
    ]

    def first_existing(paths):
        for p in paths:
            if p and os.path.exists(p):
                return p
        return None

    chrome_path = first_existing(chrome_candidates)
    if chrome_path:
        options.binary_location = chrome_path

    try:
        # Se esiste un chromedriver di sistema, usalo
        sys_driver = first_existing(driver_candidates)
        if sys_driver:
            service = Service(executable_path=sys_driver)
            return webdriver.Chrome(service=service, options=options)
        # Altrimenti lascia che Selenium Manager o WebDriverManager facciano il loro lavoro
        # (Selenium Manager prova da solo se non forniamo Service)
        return webdriver.Chrome(options=options)

    except (SessionNotCreatedException, WebDriverException):
        # 2) fallback: scarica il driver giusto a runtime
        service = Service(ChromeDriverManager().install())
        return webdriver.Chrome(service=service, options=options)




# ---------- Route ----------
def borsa():
    if request.method == "GET":
        return render_template("borsa.html")

    tipologia = request.form.get("tipologia")
    investimento = request.form.get("investimento")
    investimento = float(investimento.replace(",", "."))


    driver = None
    try:
        driver = make_driver()

        tutti_dati = []
        # Pagine BTP su Borsa Italiana
        for pagina in range(1, 9):
            url = f"https://www.borsaitaliana.it/borsa/obbligazioni/mot/btp/lista.html?&page={pagina}"
            driver.get(url)

            html = driver.page_source
            soup = BeautifulSoup(html, "lxml")
            for tr in soup.select("table tbody tr"):
                celle = [td.get_text(strip=True) for td in tr.find_all("td")]
                if celle:
                    tutti_dati.append(celle)



        # Colonne attese dalla tabella corrente
        colonne = ["ISIN", "Nome", "Prezzo", "Cedola (%)", "Scadenza", "Altro"]
        df = pd.DataFrame(tutti_dati, columns=colonne)

        # Tipi
        df["ISIN"] = df["ISIN"].astype(str).str.strip().str[0:12]

        df["Nome"] = df["Nome"].astype(str)

        df["Cedola (%)"] = (
            df["Cedola (%)"]
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df["Cedola (%)"] = pd.to_numeric(df["Cedola (%)"], errors="coerce")
        pd.set_option("display.float_format", lambda x: f"{x:.5f}")

        df["Scadenza"] = pd.to_datetime(df["Scadenza"], dayfirst=True, errors="coerce").dt.normalize()
        oggi = pd.to_datetime(datetime.today()).normalize()
        df["GG"] = (df["Scadenza"] - oggi).dt.days
        df["Scadenza"] = df["Scadenza"].dt.strftime("%d-%m-%y")

        # Prezzo e Cedola: da "113.500" o "113,50" a float
        df["Prezzo"] = (
            df["Prezzo"]
            .str.replace(".", "", regex=False)   # rimuove puntini come mille/format
            .str.replace(",", ".", regex=False)  # virgola -> punto decimale
        )
        df["Prezzo"] = pd.to_numeric(df["Prezzo"], errors="coerce")

        
        df["ValNOM (€)"] = investimento * 100 / df["Prezzo"]
        df["Rend. (%)"] = df["Cedola (%)"] / df["Prezzo"] * 100 * 2
        df["CedLRD (€)"] = df["Rend. (%)"] / 100 * investimento / 2
        df["CedNTT (€)"] = df["CedLRD (€)"] * 0.875
        df["Ced12m (%)"] = df["Cedola (%)"] * 2
        df["RateoLRD (€)"] =  df["CedLRD (€)"]  -  (((df["GG"] % 182.5)/182.5) * df["CedLRD (€)"])
        df["RateoNTT (€)"] = df["RateoLRD (€)"] * 0.875
        df["EsbINIZ"] = investimento + df["RateoNTT (€)"]
        df["CedNUM"] = df["GG"] // 182.5 + 1
        df["RicLRD12m (€)"] = df["CedLRD (€)"] * 2
        df["RicNTT12M (€)"] = df["CedLRD (€)"] * 2 * 0.875
        df["RicTOT (€)"] = (df["CedLRD (€)"] * df["CedNUM"]) + df["ValNOM (€)"]
        df["GuadTOTLRD (€)"] = df["RicTOT (€)"] - df["EsbINIZ"]
        df["GuadTOTNTT (€)"] = df["GuadTOTLRD (€)"] * 0.875
        df["GuadLRD12m (€)"] = df["GuadTOTLRD (€)"] / df["GG"] * 365
        df["GuadNTT12m (€)"] = df["GuadLRD12m (€)"] * 0.875

        # Elimina "Altro" se presente
        if "Altro" in df.columns:
            df = df.drop(columns=["Altro"])

        colonne_da_esportare = [
            "ISIN",
            "Nome",
            "Prezzo",
            "Cedola (%)",
            "Scadenza",
        ]

        # Salvataggio XLSX: preferisci /tmp se scrivibile (Render)
        output_path = "/tmp/btp_dati.xlsx" if os.access("/tmp", os.W_OK) else "btp_dati.xlsx"
        df[colonne_da_esportare].to_excel(output_path, index=False)

        # Auto-larghezza colonne
        wb = load_workbook(output_path)
        ws = wb.active
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(12, int(max_length * 1.2))
        wb.save(output_path)

        df.replace([np.inf, -np.inf], np.nan, inplace=True)
        df = df.fillna('')

        if "ISIN" in df.columns:
            def isin_link(x):
                if pd.isna(x) or str(x).strip() == "":
                    return ""
                s = str(x).strip().upper()
                return (
                    f'<a href="https://www.borsaitaliana.it/borsa/obbligazioni/mot/btp/scheda/{s}.html?lang=it" '
                    f'target="_blank" rel="noopener noreferrer">{s}</a>'
                )

            df["ISIN"] = df["ISIN"].map(isin_link)

        def aggiungi_info(nome):
            if not isinstance(nome, str):
                return nome
            info = None
            if nome.lower().startswith("Btp Coupon"):
                info = (
                    "Indicizzato all'inflazione: NO&#10;"
                    "Premio finale: NO&#10;"
                    "Periodicità cedola: NO&#10;"
                    "Cedola fissa/variabile: NO&#10;"
                    "Note particolari: meccanismo del coupon stripping &#10"
                )
            elif nome.lower().startswith("btp coupon"):
                info = "BTP coupon"
            if info:
                return (
                    f'{nome} '
                    f'<span title="{info}" '
                    f'style="display:inline-block; width:16px; height:16px; '
                    f'border-radius:50%; background-color:#4CAF50; color:#fff; '
                    f'text-align:center; font-size:12px; line-height:16px; '
                    f'cursor:help; font-weight:bold;">i</span>'
                )
            return nome

        df["Nome"] = df["Nome"].apply(aggiungi_info)

        # Tabella HTML (id per DataTables)
        tabella_html = df.to_html(
            classes="tabella-risultati display nowrap",
            table_id="btpTable",
            index=False,
            escape=False
        )

        return render_template(
            "borsa.html",
            risultato=f"Elaborazione {tipologia} completata. File salvato come {os.path.basename(output_path)}",
            tabella=tabella_html,
        )

    except Exception as e:
        logging.exception("Errore durante scraping/elaborazione BTP")
        # Mostra l'errore completo anche a video per debug
        return render_template(
            "borsa.html",
            risultato=f"Si è verificato un errore: {type(e).__name__} - {e}",
            tabella=None
        ), 500


    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass
